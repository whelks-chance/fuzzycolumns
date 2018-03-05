import numbers
import pprint
import sys

import os
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
from openpyxl.cell import Cell
from fuzzywuzzy import fuzz


class Bucket:
    def __init__(self, lower, upper):
        assert isinstance(lower, numbers.Real)
        assert isinstance(upper, numbers.Real)
        assert upper >= lower
        self.lower = lower
        self.upper = upper
        self.matches = []

    def __getitem__(self, item):
        return self.__dict__()

    def __dict__(self):
        return {
            'upper': self.upper,
            'lower': self.lower,
            'matches': self.matches_detailed()
        }

    def __repr__(self):
        return str(self.__dict__())

    def record(self, wb1_cell, wb2_cell, ratio):
        self.matches.append(
            {
                'wb1_cell': wb1_cell,
                'wb2_cell': wb2_cell,
                'ratio': ratio
            }
        )

    def matches_detailed(self, detail=0):
        all_data = []
        for m in self.matches:
            all_data.append({
                'ratio': m['ratio'],
                'cell1': {
                    'value': m['wb1_cell'].value,
                    'col_idx': m['wb1_cell'].col_idx,
                    'column': m['wb1_cell'].column,
                    'sheet': m['wb1_cell'].parent.title,
                    'coordinate': m['wb1_cell'].coordinate,
                    'row': m['wb1_cell'].row
                },
                'cell2': {
                    'value': m['wb2_cell'].value,
                    'col_idx': m['wb2_cell'].col_idx,
                    'column': m['wb2_cell'].column,
                    'sheet': m['wb2_cell'].parent.title,
                    'coordinate': m['wb2_cell'].coordinate,
                    'row': m['wb2_cell'].row
                }
            })

        return all_data


class FuzzyReport:
    def __init__(self, levels=None):
        if levels:
            self.levels = levels
        else:
            self.levels = [30, 60, 80]
    # So ranges are 0-30, 30-60, 60-80, 80-100
    # number of ranges is len(levels) + 1

        previous_level = 0
        self.buckets = []
        for l in self.levels:
            self.buckets.append(Bucket(previous_level, l))
            previous_level = l
        self.buckets.append(Bucket(previous_level, 100))

    def __getitem__(self, item):
        print('*', item, '*')
        return self.__dict__()

    def number_of_buckets(self):
        return len(self.buckets)

    def get_bucket_by_ratio(self, ratio):
        for b in self.buckets:
            if b.lower < ratio <= b.upper:
                return b

    def record(self, wb1_cell, wb2_cell, ratio):
        assert isinstance(wb1_cell, Cell)
        assert isinstance(wb2_cell, Cell)

        # print('{} : {} : {}'.format(wb1_cell.value, wb2_cell.value, ratio))

        bucket = self.get_bucket_by_ratio(ratio)
        bucket.record(wb1_cell, wb2_cell, ratio)

    def __repr__(self):
        return pprint.pformat(self.buckets)

    def __dict__(self):
        return self.buckets

    def print_buckets(self):
        for idx, b in enumerate(self.buckets):
            self.print_bucket(idx)

    def print_bucket(self, bucket_idx, detail=0):
        if bucket_idx >= len(self.buckets):
            bucket_idx = len(self.buckets) -1
        b = self.buckets[bucket_idx]
        assert isinstance(b, Bucket)
        print('{} : {} : {}'.format(b.lower, b.upper, b.matches_detailed(detail=detail)))

    def print_distribution(self):
        for b in self.buckets:
            print('{} {} {}'.format(b.lower, b.upper, len(b.matches)))

    def best_matches(self):
        for i in reversed(range(0, len(self.buckets) - 1)):
            # print(i)
            # print(self.buckets[i])
            if len(self.buckets[i]['matches']):
                return self.buckets[i]


class FuzzyColumns:
    def __init__(self):
        pass

    def load_wb(self, file):
        try:
            return load_workbook(filename=file)
        except InvalidFileException as ife:
            file_format = os.path.splitext(file)[-1]
            print('Found an {} which failed to be read.'.format(file_format))
            print(ife)
            raise ife
        except Exception as e:
            raise e

    def compare_spreadsheets(self, file1, file2, levels=None):
        try:
            wb1 = self.load_wb(file1)
            wb2 = self.load_wb(file2)
        except Exception as e:
            raise e

        return self.compare_workbooks(wb1, wb2, levels=levels)

    def compare_workbooks(self, wb1, wb2, levels=None):
        report = FuzzyReport(levels=levels)

        for sheet_name in wb1.sheetnames:
            sheet = wb1[sheet_name]

            # print('WB {} Sheet {}'.format(wb1.code_name, sheet_name))
            # print('Dimensions {}'.format(sheet.dimensions))

            for row in list(sheet.rows)[:1]:
                for cell in row:

                    wb1_cell = cell.value.lower()

                    for sheet_name in wb2.sheetnames:
                        sheet = wb2[sheet_name]

                        # print('WB {} Sheet {}'.format(wb2.code_name, sheet_name))
                        # print('Dimensions {}'.format(sheet.dimensions))

                        for row2 in list(sheet.rows)[:1]:
                            for cell2 in row2:
                                if cell2.value is not None:
                                    wb2_cell = cell2.value.lower()

                                    ratio = fuzz.token_sort_ratio(
                                        wb1_cell,
                                        wb2_cell
                                    )
                                    report.record(cell, cell2, ratio)
        return report


if __name__ == '__main__':
    args = sys.argv

    fc = FuzzyColumns()
    if len(args) > 2:
        report = fc.compare_spreadsheets(args[1], args[2], levels=[5, 30, 70, 80, 90, 95])
        # report.print_buckets()
        report.print_bucket(4, detail=0)
        print(report.number_of_buckets())
        report.print_distribution()

    else:
        print("Need two files to compare")